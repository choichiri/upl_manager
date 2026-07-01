"""엑셀 파일 읽기/쓰기 핸들러"""
import shutil
from pathlib import Path
from datetime import datetime
import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.styles import Alignment

from src.models import Customer, ProgressEntry


# 시스템 시트 (고객사 시트가 아닌 것들)
SYSTEM_SHEETS_KOREAN = {"Index", "검색"}
SYSTEM_SHEETS_GLOBAL = {"Index", "검색", "목록"}


class ExcelHandler:
    """엑셀 파일에서 고객사 데이터를 읽고 쓰는 핸들러"""

    def __init__(self, file_path: str, file_type: str = "korean"):
        self.file_path = Path(file_path)
        self.file_type = file_type
        self.system_sheets = (
            SYSTEM_SHEETS_KOREAN if file_type == "korean"
            else SYSTEM_SHEETS_GLOBAL
        )
        self._wb: Workbook | None = None

    def load(self) -> list[Customer]:
        """엑셀 파일을 읽어 고객사 목록을 반환한다."""
        if not self.file_path.exists():
            raise FileNotFoundError(f"파일을 찾을 수 없습니다: {self.file_path}")

        self._wb = openpyxl.load_workbook(
            str(self.file_path), keep_vba=True
        )

        customers = []
        for sheet_name in self._wb.sheetnames:
            if sheet_name in self.system_sheets:
                continue
            customer = self._read_customer_sheet(sheet_name)
            if customer:
                customers.append(customer)

        return customers

    def reload(self) -> list[Customer]:
        """파일을 다시 로드한다."""
        self.close()
        return self.load()

    def _read_customer_sheet(self, sheet_name: str) -> Customer | None:
        """개별 고객사 시트에서 데이터를 읽는다."""
        ws = self._wb[sheet_name]
        customer = Customer(sheet_name=sheet_name)

        customer.company_name = self._cell_str(ws, "A3")
        customer.address_or_country = self._cell_str(ws, "C3")
        customer.ceo = self._cell_str(ws, "E3")
        customer.business_number = self._cell_str(ws, "A5")
        customer.contact_person = self._cell_str(ws, "E4")
        customer.phone = self._cell_str(ws, "E5")
        customer.email = self._cell_str(ws, "E6")
        customer.progress = self._read_progress(ws)

        return customer

    def _read_progress(self, ws) -> list[ProgressEntry]:
        """진행사항 목록을 읽는다. 최신순(역순)으로 반환."""
        entries = []
        for row_num in range(10, ws.max_row + 1):
            date_val = ws.cell(row=row_num, column=1).value
            content_val = ws.cell(row=row_num, column=2).value

            if date_val is None and content_val is None:
                continue

            content = str(content_val).strip() if content_val else ""
            if not content:
                continue

            # 날짜(A열)가 비어있는 행은 직전 항목의 '이어지는 줄'로 병합
            if date_val is None and entries:
                entries[-1].content += "\n" + content
                continue

            entries.append(ProgressEntry(date=date_val, content=content))

        entries.reverse()
        return entries

    # === 쓰기 기능 ===

    def delete_customer(self, customer: Customer):
        """고객사 시트를 삭제하고 Index에서도 제거"""
        self._check_file_writable()
        self._reload_workbook()

        # 시트 삭제
        if customer.sheet_name in self._wb.sheetnames:
            del self._wb[customer.sheet_name]

        # Index에서 해당 행 클리어
        if "Index" in self._wb.sheetnames:
            ws = self._wb["Index"]
            max_col = ws.max_column or 8
            for row_num in range(2, ws.max_row + 1):
                if ws.cell(row=row_num, column=1).value == customer.sheet_name:
                    for col in range(1, max_col + 1):
                        ws.cell(row=row_num, column=col).value = ""
                    break

        self._save_file()

    def save_customer_info(self, customer: Customer):
        """고객사 기본정보를 시트에 저장"""
        self._check_file_writable()
        self._reload_workbook()

        ws = self._wb[customer.sheet_name]
        ws["A3"] = customer.company_name
        ws["C3"] = customer.address_or_country
        ws["E3"] = customer.ceo
        ws["A5"] = customer.business_number
        ws["E4"] = customer.contact_person
        ws["E5"] = customer.phone
        ws["E6"] = customer.email

        # Index 시트도 업데이트
        self._update_index(customer)
        self._save_file()

    def _find_last_data_row(self, ws, start_row=10) -> int:
        """실제 데이터가 있는 마지막 행 번호를 반환"""
        last = start_row - 1
        for row_num in range(start_row, ws.max_row + 1):
            if ws.cell(row=row_num, column=1).value or ws.cell(row=row_num, column=2).value:
                last = row_num
        return last

    def _merge_content_cells(self, ws, row):
        """B~E 셀 병합 (이미 병합되어 있으면 무시)"""
        from openpyxl.utils import get_column_letter
        merge_range = f"B{row}:E{row}"
        # 이미 병합되어 있는지 확인
        for m in ws.merged_cells.ranges:
            if str(m) == merge_range:
                return
        ws.merge_cells(merge_range)

    def add_progress_entry(self, customer: Customer, entry: ProgressEntry):
        """진행사항을 시트 맨 아래에 추가하고 파일 저장"""
        self._check_file_writable()
        self._reload_workbook()

        ws = self._wb[customer.sheet_name]

        next_row = self._find_last_data_row(ws) + 1
        cell = ws.cell(row=next_row, column=1, value=entry.date)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=next_row, column=2, value=entry.content)
        self._merge_content_cells(ws, next_row)

        self._save_file()

    def save_all_progress(self, customer: Customer):
        """고객사의 전체 진행사항을 시트에 덮어쓰기 (수정/삭제 반영)"""
        self._check_file_writable()
        self._reload_workbook()

        ws = self._wb[customer.sheet_name]

        # 기존 진행사항 전부 클리어 (빈 문자열로 — None은 xlsm에서 무시됨)
        for row_num in range(10, ws.max_row + 1):
            for col in (1, 2):
                cell = ws.cell(row=row_num, column=col)
                if not isinstance(cell, openpyxl.cell.cell.MergedCell):
                    cell.value = ""

        # 메모리의 customer.progress를 기준으로 다시 저장 (최신순 → 오래된순)
        entries = list(reversed(customer.progress))
        for i, entry in enumerate(entries):
            row = 10 + i
            a_cell = ws.cell(row=row, column=1)
            b_cell = ws.cell(row=row, column=2)
            if not isinstance(a_cell, openpyxl.cell.cell.MergedCell):
                a_cell.value = entry.date
                a_cell.alignment = Alignment(horizontal="center", vertical="center")
            if not isinstance(b_cell, openpyxl.cell.cell.MergedCell):
                b_cell.value = entry.content

        self._save_file()

    def _find_template_sheet(self):
        """복사할 템플릿 시트를 찾는다 (시스템 시트 제외 마지막 고객사 시트)"""
        for name in reversed(self._wb.sheetnames):
            if name not in self.system_sheets:
                return self._wb[name]
        return None

    def create_customer_sheet(self, customer: Customer):
        """기존 고객사 시트를 복사하여 새 시트를 생성하고 데이터만 교체한다."""
        self._check_file_writable()
        self._reload_workbook()

        template = self._find_template_sheet()

        if template:
            # 기존 시트 복사 (서식, 병합, 열너비 등 유지)
            ws = self._wb.copy_worksheet(template)
            ws.title = customer.sheet_name

            # 진행사항 셀 클리어 (10행부터)
            for row_num in range(10, ws.max_row + 1):
                for col in (1, 2):
                    cell = ws.cell(row=row_num, column=col)
                    if not isinstance(cell, openpyxl.cell.cell.MergedCell):
                        cell.value = ""
        else:
            # 복사할 시트가 없으면 직접 생성 (fallback)
            ws = self._wb.create_sheet(title=customer.sheet_name)
            ws["A1"] = "영 업 일 지"
            ws["A2"] = "고객사명"
            ws["E2"] = "대표자"
            if self.file_type == "korean":
                ws["C2"] = "주  소"
            else:
                ws["C2"] = "국 가"
            ws["A4"] = "사업자등록번호"
            ws["C4"] = "담당자"
            ws["D4"] = "성명/직급"
            ws["D5"] = "연락처"
            ws["A6"] = "고객코드"
            ws["D6"] = "Email"
            ws["A8"] = "진행 사항"
            ws["A9"] = "일 자"
            ws["B9"] = "내    용"

        # 데이터 셀만 새 값으로 교체
        ws["A3"] = customer.company_name
        ws["C3"] = customer.address_or_country
        ws["E3"] = customer.ceo
        ws["A5"] = customer.business_number
        ws["E4"] = customer.contact_person
        ws["E5"] = customer.phone
        ws["E6"] = customer.email

        # Index 업데이트
        self._update_index(customer)
        self._save_file()

    def _update_index(self, customer: Customer):
        """Index 시트에 고객사 정보를 추가/업데이트"""
        if "Index" not in self._wb.sheetnames:
            return

        ws = self._wb["Index"]

        # 기존 행 찾기
        existing_row = None
        for row_num in range(2, ws.max_row + 1):
            if ws.cell(row=row_num, column=1).value == customer.sheet_name:
                existing_row = row_num
                break

        target_row = existing_row or (ws.max_row + 1)

        ws.cell(row=target_row, column=1, value=customer.sheet_name)

        if self.file_type == "global":
            # 해외: A=시트명, B=국가, C=고객사명, D=대표자, E=사업자번호, F=담당자, G=연락처, H=이메일
            ws.cell(row=target_row, column=2, value=customer.address_or_country)
            ws.cell(row=target_row, column=3, value=customer.company_name)
            ws.cell(row=target_row, column=4, value=customer.ceo)
            ws.cell(row=target_row, column=5, value=customer.business_number)
            ws.cell(row=target_row, column=6, value=customer.contact_person)
            ws.cell(row=target_row, column=7, value=customer.phone)
            ws.cell(row=target_row, column=8, value=customer.email)
        else:
            # 국내: A=시트명, B=고객사명, C=대표자, D=사업자번호, E=담당자, F=연락처, G=이메일
            ws.cell(row=target_row, column=2, value=customer.company_name)
            ws.cell(row=target_row, column=3, value=customer.ceo)
            ws.cell(row=target_row, column=4, value=customer.business_number)
            ws.cell(row=target_row, column=5, value=customer.contact_person)
            ws.cell(row=target_row, column=6, value=customer.phone)
            ws.cell(row=target_row, column=7, value=customer.email)

    def backup(self):
        """저장 전 백업 파일 생성"""
        if not self.file_path.exists():
            return
        backup_dir = self.file_path.parent / "_백업"
        backup_dir.mkdir(exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        stem = self.file_path.stem
        suffix = self.file_path.suffix
        backup_path = backup_dir / f"{stem}_{timestamp}{suffix}"
        shutil.copy2(str(self.file_path), str(backup_path))

    def _check_file_writable(self):
        """파일이 쓰기 가능한지 먼저 확인"""
        try:
            with open(str(self.file_path), "a"):
                pass
        except (PermissionError, OSError):
            raise PermissionError(
                "엑셀 파일이 다른 프로그램에서 열려 있어 저장할 수 없습니다.\n"
                "Excel을 닫은 후 다시 시도해주세요."
            )

    def _reload_workbook(self):
        """디스크에서 워크북을 다시 로드 (엑셀이 파일을 변경했을 수 있으므로)"""
        if self._wb:
            try:
                self._wb.close()
            except Exception:
                pass
        self._wb = openpyxl.load_workbook(
            str(self.file_path), keep_vba=True
        )

    def _save_file(self):
        """워크북을 파일로 저장"""
        self._check_file_writable()
        self._wb.save(str(self.file_path))

    @staticmethod
    def _cell_str(ws, coord: str) -> str:
        val = ws[coord].value
        if val is None:
            return ""
        return str(val).strip()

    @property
    def sheet_count(self) -> int:
        if self._wb is None:
            return 0
        return len([s for s in self._wb.sheetnames if s not in self.system_sheets])

    def close(self):
        if self._wb:
            self._wb.close()
            self._wb = None
